from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule

wb = Workbook()
ws = wb.active
ws.title = "Carteira"

DARK_BG   = "1E1E2E"
HEADER_BG = "2A2D3E"
GREEN_BG  = "1A3A2A"
RED_BG    = "3A1A1A"
GREEN_TXT = "00C896"
RED_TXT   = "FF5C5C"
WHITE     = "FFFFFF"
LIGHT_GRAY= "C0C0C0"
YELLOW    = "FFD700"
BLUE_TXT  = "5B9BD5"
SUBHDR_BG = "343650"
ORANGE    = "FFA500"

thin = Side(style='thin', color="404060")
med  = Side(style='medium', color="5B5B8A")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_med  = Border(left=med,  right=med,  top=med,  bottom=med)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def right_align():
    return Alignment(horizontal='right', vertical='center')

FMT_DATE = 'DD/MM/YYYY'
FMT_CURR = 'R$ #,##0.00;(R$ #,##0.00);"-"'
FMT_PCT  = '0.00%;(0.00%);"-"'

# ── Title ─────────────────────────────────────────────────────────────────────
ws.merge_cells('A1:L1')
ws['A1'] = "CONTROLE DE CARTEIRA DE INVESTIMENTOS"
ws['A1'].font      = Font(name='Arial', bold=True, color=YELLOW, size=14)
ws['A1'].fill      = fill(DARK_BG)
ws['A1'].alignment = center()
ws.row_dimensions[1].height = 32

# ── Column headers row 2 ──────────────────────────────────────────────────────
headers = [
    ("A", "ATIVO",           12),
    ("B", "DATA\nCOMPRA",    13),
    ("C", "PRECO\nCOMPRA",   13),
    ("D", "QTD",              8),
    ("E", "TOTAL\nINVESTIDO",14),
    ("F", "DATA\nVENDA",     13),
    ("G", "PRECO\nVENDA",    13),
    ("H", "TOTAL\nVENDA",    14),
    ("I", "RESULTADO\nR$",   15),
    ("J", "RESULTADO\n%",    13),
    ("K", "STATUS",          11),
]

for col, title, width in headers:
    c = ws[f"{col}2"]
    c.value     = title
    c.font      = Font(name='Arial', bold=True, color=YELLOW, size=9)
    c.fill      = fill(HEADER_BG)
    c.alignment = center()
    c.border    = border_med
    ws.column_dimensions[col].width = width

ws.row_dimensions[2].height = 30

# ── Sample data rows 3-22 ────────────────────────────────────────────────────
sample = [
    ("MU",   "2026-01-15", 865.00, 10, "2026-03-20", 920.00),
    ("AAPL", "2026-01-15", 220.50,  5, "2026-03-20", 240.00),
    ("TSLA", "2026-02-10", 180.00,  8, "2026-04-05", 155.00),
    ("NVDA", "2026-02-10", 850.00,  3, None,         None),
    ("MSFT", "2026-03-01", 380.00,  4, None,         None),
]

DATA_START = 3
DATA_END   = 22

for i, row_data in enumerate(sample, start=DATA_START):
    r = i
    ativo, dc, pc, qtd, dv, pv = row_data
    row_fill = fill("252840") if r % 2 == 0 else fill("1E2035")

    ws[f'A{r}'] = ativo
    ws[f'A{r}'].font = Font(name='Arial', bold=True, color=YELLOW, size=10)

    ws[f'B{r}'] = dc
    ws[f'B{r}'].number_format = FMT_DATE

    ws[f'C{r}'] = pc
    ws[f'C{r}'].number_format = FMT_CURR

    ws[f'D{r}'] = qtd

    ws[f'E{r}'] = f'=IF(C{r}<>"",C{r}*D{r},"")'
    ws[f'E{r}'].number_format = FMT_CURR

    if dv:
        ws[f'F{r}'] = dv
        ws[f'F{r}'].number_format = FMT_DATE

    if pv:
        ws[f'G{r}'] = pv
    ws[f'G{r}'].number_format = FMT_CURR

    ws[f'H{r}'] = f'=IF(G{r}<>"",G{r}*D{r},"")'
    ws[f'H{r}'].number_format = FMT_CURR

    ws[f'I{r}'] = f'=IF(H{r}<>"",H{r}-E{r},"")'
    ws[f'I{r}'].number_format = FMT_CURR

    ws[f'J{r}'] = f'=IF(AND(H{r}<>"",E{r}<>0),(H{r}-E{r})/E{r},"")'
    ws[f'J{r}'].number_format = FMT_PCT

    ws[f'K{r}'] = f'=IF(F{r}<>"","Fechada","Aberta")'

    for col in 'ABCDEFGHIJK':
        c = ws[f'{col}{r}']
        c.fill   = row_fill
        c.border = border_thin
        c.alignment = center() if col in ('A','B','F','K') else right_align()
        if c.font.name == 'Calibri':
            c.font = Font(name='Arial', color=WHITE, size=10)

    ws.row_dimensions[r].height = 20

# Empty rows
for r in range(DATA_START + len(sample), DATA_END + 1):
    row_fill = fill("252840") if r % 2 == 0 else fill("1E2035")
    for col in 'ABCDEFGHIJK':
        c = ws[f'{col}{r}']
        c.fill      = row_fill
        c.border    = border_thin
        c.font      = Font(name='Arial', color=WHITE, size=10)
        c.alignment = center() if col in ('A','B','F','K') else right_align()

    ws[f'E{r}'] = f'=IF(C{r}<>"",C{r}*D{r},"")'
    ws[f'E{r}'].number_format = FMT_CURR
    ws[f'G{r}'].number_format = FMT_CURR
    ws[f'H{r}'] = f'=IF(G{r}<>"",G{r}*D{r},"")'
    ws[f'H{r}'].number_format = FMT_CURR
    ws[f'I{r}'] = f'=IF(H{r}<>"",H{r}-E{r},"")'
    ws[f'I{r}'].number_format = FMT_CURR
    ws[f'J{r}'] = f'=IF(AND(H{r}<>"",E{r}<>0),(H{r}-E{r})/E{r},"")'
    ws[f'J{r}'].number_format = FMT_PCT
    ws[f'K{r}'] = f'=IF(F{r}<>"","Fechada","Aberta")'
    ws.row_dimensions[r].height = 20

# ── SECTION: Resultado por Data de Compra ─────────────────────────────────────
SEC1 = DATA_END + 2  # row 24

ws.merge_cells(f'A{SEC1}:K{SEC1}')
ws[f'A{SEC1}'] = "RESULTADO POR DATA DE COMPRA"
ws[f'A{SEC1}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=11)
ws[f'A{SEC1}'].fill      = fill(YELLOW)
ws[f'A{SEC1}'].alignment = center()
ws.row_dimensions[SEC1].height = 24

# Sub-headers for date breakdown
SEC1H = SEC1 + 1
sub_hdrs = [
    ("B", "DATA COMPRA"),
    ("E", "TOTAL INVEST."),
    ("H", "TOTAL VENDA"),
    ("I", "RESULTADO R$"),
    ("J", "RESULTADO %"),
]
for col, title in sub_hdrs:
    c = ws[f'{col}{SEC1H}']
    c.value     = title
    c.font      = Font(name='Arial', bold=True, color=YELLOW, size=9)
    c.fill      = fill(HEADER_BG)
    c.alignment = center()
    c.border    = border_med

for col in 'ABCDEFGHIJK':
    c = ws[f'{col}{SEC1H}']
    c.fill   = fill(HEADER_BG)
    c.border = border_med
ws.row_dimensions[SEC1H].height = 22

# Date breakdown rows — one per unique purchase date used in sample
dates_used = ["2026-01-15", "2026-02-10", "2026-03-01"]
date_labels = ["15/01/2026", "10/02/2026", "01/03/2026"]

date_rows = []
for idx, (dt, lbl) in enumerate(zip(dates_used, date_labels)):
    r = SEC1H + 1 + idx
    date_rows.append(r)
    row_fill = fill("252840") if idx % 2 == 0 else fill("1E2035")

    ws[f'B{r}'] = lbl
    ws[f'B{r}'].font      = Font(name='Arial', bold=True, color=YELLOW, size=10)
    ws[f'B{r}'].alignment = center()
    ws[f'B{r}'].fill      = row_fill
    ws[f'B{r}'].border    = border_thin

    # Total investido nessa data
    ws[f'E{r}'] = f'=SUMPRODUCT((TEXT(B{DATA_START}:B{DATA_END},"DD/MM/YYYY")="{lbl}")*IFERROR(E{DATA_START}:E{DATA_END},0))'
    ws[f'E{r}'].number_format = FMT_CURR
    ws[f'E{r}'].font      = Font(name='Arial', color=BLUE_TXT, size=10, bold=True)
    ws[f'E{r}'].alignment = right_align()
    ws[f'E{r}'].fill      = row_fill
    ws[f'E{r}'].border    = border_thin

    # Total venda nessa data
    ws[f'H{r}'] = f'=SUMPRODUCT((TEXT(B{DATA_START}:B{DATA_END},"DD/MM/YYYY")="{lbl}")*IFERROR(H{DATA_START}:H{DATA_END},0))'
    ws[f'H{r}'].number_format = FMT_CURR
    ws[f'H{r}'].font      = Font(name='Arial', color=WHITE, size=10)
    ws[f'H{r}'].alignment = right_align()
    ws[f'H{r}'].fill      = row_fill
    ws[f'H{r}'].border    = border_thin

    # Resultado R$ nessa data
    ws[f'I{r}'] = f'=IF(H{r}<>0,H{r}-E{r},"")'
    ws[f'I{r}'].number_format = FMT_CURR
    ws[f'I{r}'].font      = Font(name='Arial', bold=True, size=10)
    ws[f'I{r}'].alignment = right_align()
    ws[f'I{r}'].fill      = row_fill
    ws[f'I{r}'].border    = border_thin

    # Resultado % nessa data
    ws[f'J{r}'] = f'=IF(AND(I{r}<>"",E{r}<>0),I{r}/E{r},"")'
    ws[f'J{r}'].number_format = FMT_PCT
    ws[f'J{r}'].font      = Font(name='Arial', bold=True, size=10)
    ws[f'J{r}'].alignment = right_align()
    ws[f'J{r}'].fill      = row_fill
    ws[f'J{r}'].border    = border_thin

    for col in 'ACDFGK':
        c = ws[f'{col}{r}']
        c.fill   = row_fill
        c.border = border_thin

    ws.row_dimensions[r].height = 20

# ── Media row ─────────────────────────────────────────────────────────────────
MEDIA_ROW = date_rows[-1] + 1
i_refs_str = ",".join([f'I{r}' for r in date_rows])
j_refs_str = ",".join([f'J{r}' for r in date_rows])
e_refs_str = ",".join([f'E{r}' for r in date_rows])

ws.merge_cells(f'A{MEDIA_ROW}:D{MEDIA_ROW}')
ws[f'A{MEDIA_ROW}'] = "MEDIA POR DATA"
ws[f'A{MEDIA_ROW}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=10)
ws[f'A{MEDIA_ROW}'].fill      = fill(ORANGE)
ws[f'A{MEDIA_ROW}'].alignment = right_align()
ws[f'A{MEDIA_ROW}'].border    = border_med

ws[f'E{MEDIA_ROW}'] = f'=AVERAGE({e_refs_str})'
ws[f'E{MEDIA_ROW}'].number_format = FMT_CURR
ws[f'E{MEDIA_ROW}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=10)
ws[f'E{MEDIA_ROW}'].fill      = fill(ORANGE)
ws[f'E{MEDIA_ROW}'].alignment = right_align()
ws[f'E{MEDIA_ROW}'].border    = border_med

ws[f'I{MEDIA_ROW}'] = f'=IFERROR(AVERAGEIF({i_refs_str},"<>"""),"")'
ws[f'I{MEDIA_ROW}'].number_format = FMT_CURR
ws[f'I{MEDIA_ROW}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=10)
ws[f'I{MEDIA_ROW}'].fill      = fill(ORANGE)
ws[f'I{MEDIA_ROW}'].alignment = right_align()
ws[f'I{MEDIA_ROW}'].border    = border_med

ws[f'J{MEDIA_ROW}'] = f'=IFERROR(AVERAGEIF({j_refs_str},"<>"""),"")'
ws[f'J{MEDIA_ROW}'].number_format = FMT_PCT
ws[f'J{MEDIA_ROW}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=10)
ws[f'J{MEDIA_ROW}'].fill      = fill(ORANGE)
ws[f'J{MEDIA_ROW}'].alignment = right_align()
ws[f'J{MEDIA_ROW}'].border    = border_med

for col in 'BDFGHK':
    c = ws[f'{col}{MEDIA_ROW}']
    c.fill   = fill(ORANGE)
    c.border = border_med
ws.row_dimensions[MEDIA_ROW].height = 24

# ── SECTION: Totais Gerais ─────────────────────────────────────────────────────
SEC2 = MEDIA_ROW + 2

ws.merge_cells(f'A{SEC2}:K{SEC2}')
ws[f'A{SEC2}'] = "TOTAIS GERAIS"
ws[f'A{SEC2}'].font      = Font(name='Arial', bold=True, color=DARK_BG, size=11)
ws[f'A{SEC2}'].fill      = fill(YELLOW)
ws[f'A{SEC2}'].alignment = center()
ws.row_dimensions[SEC2].height = 24

def total_row(r, label, formula, fmt, color):
    ws.merge_cells(f'A{r}:H{r}')
    ws[f'A{r}'] = label
    ws[f'A{r}'].font      = Font(name='Arial', bold=True, color=color, size=10)
    ws[f'A{r}'].fill      = fill(SUBHDR_BG)
    ws[f'A{r}'].alignment = right_align()
    ws[f'A{r}'].border    = border_thin

    ws[f'I{r}'] = formula
    ws[f'I{r}'].number_format = fmt
    ws[f'I{r}'].font      = Font(name='Arial', bold=True, color=color, size=11)
    ws[f'I{r}'].fill      = fill(SUBHDR_BG)
    ws[f'I{r}'].alignment = right_align()
    ws[f'I{r}'].border    = border_thin

    for col in ('J', 'K'):
        ws[f'{col}{r}'].fill   = fill(SUBHDR_BG)
        ws[f'{col}{r}'].border = border_thin
    ws.row_dimensions[r].height = 24

total_row(SEC2+1, "Total Investido:",
    f'=SUMPRODUCT((C{DATA_START}:C{DATA_END}<>"")*IFERROR(E{DATA_START}:E{DATA_END},0))',
    FMT_CURR, BLUE_TXT)
total_row(SEC2+2, "Total Lucro (operacoes positivas):",
    f'=SUMPRODUCT((IFERROR(I{DATA_START}:I{DATA_END},0)>0)*IFERROR(I{DATA_START}:I{DATA_END},0))',
    FMT_CURR, GREEN_TXT)
total_row(SEC2+3, "Total Perda (operacoes negativas):",
    f'=SUMPRODUCT((IFERROR(I{DATA_START}:I{DATA_END},0)<0)*IFERROR(I{DATA_START}:I{DATA_END},0))',
    FMT_CURR, RED_TXT)
total_row(SEC2+4, "Resultado Liquido:",
    f'=SUMPRODUCT(IFERROR(I{DATA_START}:I{DATA_END},0))',
    FMT_CURR, YELLOW)
total_row(SEC2+5, "Retorno Total %:",
    f'=IFERROR(I{SEC2+4}/I{SEC2+1},"")',
    FMT_PCT, YELLOW)

# ── Conditional formatting ─────────────────────────────────────────────────────
green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
red_fill   = PatternFill(start_color=RED_BG,   end_color=RED_BG,   fill_type="solid")
green_font = Font(name='Arial', bold=True, color=GREEN_TXT, size=10)
red_font   = Font(name='Arial', bold=True, color=RED_TXT,   size=10)

for rng in (f'I{DATA_START}:I{DATA_END}', f'J{DATA_START}:J{DATA_END}'):
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='lessThan',    formula=['0'], fill=red_fill,   font=red_font))

# Resultado por data — color
i_range = f'I{date_rows[0]}:I{date_rows[-1]}'
j_range = f'J{date_rows[0]}:J{date_rows[-1]}'
for rng in (i_range, j_range):
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='lessThan',    formula=['0'], fill=red_fill,   font=red_font))

ws.conditional_formatting.add(f'K{DATA_START}:K{DATA_END}',
    FormulaRule(formula=[f'=K{DATA_START}="Fechada"'],
        fill=PatternFill(start_color="1A2A3A", end_color="1A2A3A", fill_type="solid"),
        font=Font(name='Arial', color=LIGHT_GRAY, size=10)))
ws.conditional_formatting.add(f'K{DATA_START}:K{DATA_END}',
    FormulaRule(formula=[f'=K{DATA_START}="Aberta"'],
        fill=green_fill,
        font=Font(name='Arial', bold=True, color=GREEN_TXT, size=10)))

ws.freeze_panes = 'A3'
ws.sheet_properties.tabColor = "00C896"

out = r'C:\Users\Jefferson\Desktop\Carteira_Investimentos.xlsx'
wb.save(out)
print("Salvo:", out)
